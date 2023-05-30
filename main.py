import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl

# Function to handle the "Register" button click event
def register_service():
    # Get the input values
    tipo_servico = service_type_combobox.get()
    nome_cliente = customer_name_entry.get()
    numero_cliente = customer_phone_entry.get()
    pontuacao = customer_score_entry.get()

    # Validate input values
    if not all([tipo_servico, nome_cliente, numero_cliente, pontuacao]):
        messagebox.showerror("Error", "Por favor, preencha todos os campos!")
        return

    # Get the price based on the selected service type
    price = service_prices[tipo_servico]

    # Save the service details or perform further processing here
    # For simplicity, let's display a success message with customer score and price
    message = f"Serviço registrado com sucesso!.\nTipo de serviço: {tipo_servico}\nPreço: {price}\nCliente: {nome_cliente}\nPontuação: {pontuacao}"
    messagebox.showinfo("Success", message)

    # Clear the input fields
    clear_input_fields()

    # Save the data to Excel
    save_data(tipo_servico, price, nome_cliente, numero_cliente, pontuacao)

def clear_input_fields():
    service_type_combobox.set('')
    customer_name_entry.delete(0, tk.END)
    customer_phone_entry.delete(0, tk.END)
    customer_score_entry.delete(0, tk.END)

def save_data(service_type, price, customer_name, customer_phone, customer_score):
    try:
        # Load the existing workbook or create a new one
        try:
            workbook = openpyxl.load_workbook("petshop_data.xlsx")
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # Select the active sheet
        sheet = workbook.active

        if sheet.max_row == 1:
            sheet.cell(row=1, column=1).value = "Tipo de Serviço"
            sheet.cell(row=1, column=2).value = "Preço"
            sheet.cell(row=1, column=3).value = "Cliente"
            sheet.cell(row=1, column=4).value = "Número"
            sheet.cell(row=1, column=5).value = "Pontuação"


        # Append the data to the sheet
        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1).value = service_type
        sheet.cell(row=next_row, column=2).value = price
        sheet.cell(row=next_row, column=3).value = customer_name
        sheet.cell(row=next_row, column=4).value = customer_phone
        sheet.cell(row=next_row, column=5).value = customer_score

        # Save the workbook
        try:
            workbook.save("petshop_data.xlsx")
            messagebox.showinfo("Success", "Dados salvos com sucesso!")
        except PermissionError:
            messagebox.showerror("Error", "Falha ao salvar dados. Por favor, feche a planilha do Excel e tente novamente!")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


# Service prices dictionary
service_prices = {
    "Banho": "R$ 35,00",
    "Tosa": "R$ 25,00",
    "Banho e Tosa": "R$ 50,00",
    "Vermifugação": "R$ 60,00"
}

# Create the main application window
root = tk.Tk()
root.title("Cadastro de Serviço - Petshop Aroa")

# Configure the overall style
style = ttk.Style()
style.theme_use("default")
style.configure("TLabel", font=("Arial", 12))
style.configure("TButton", font=("Arial", 12))
style.configure("TEntry", font=("Arial", 12))

# Create a frame for the content
content_frame = ttk.Frame(root, padding=(20, 20))
content_frame.pack(fill=tk.BOTH, expand=True)

# Service Type
# Define service_types variable
service_types = list(service_prices.keys())

# Service Type Combobox
service_type_combobox = ttk.Combobox(content_frame, values=service_types, state="readonly")
service_type_combobox.grid(row=0, column=1, padx=10, pady=5)
service_type_label = ttk.Label(content_frame, text="Serviço:")
service_type_label.grid(row=0, column=0, sticky=tk.E)
service_type_combobox.grid(row=0, column=1, padx=10, pady=5)


price_label = ttk.Label(content_frame, text="Preço:")
price_label.grid(row=1, column=0, sticky=tk.E)
price_value = tk.StringVar()
price_value.set(service_prices[service_types[0]])
price_entry = ttk.Entry(content_frame, width=30, textvariable=price_value, state="readonly")
price_entry.grid(row=1, column=1, padx=10, pady=5)

customer_name_label = ttk.Label(content_frame, text="Nome:")
customer_name_label.grid(row=2, column=0, sticky=tk.E)
customer_name_entry = ttk.Entry(content_frame, width=30)
customer_name_entry.grid(row=2, column=1, padx=10, pady=5)


customer_phone_label = ttk.Label(content_frame, text="Número:")
customer_phone_label.grid(row=3, column=0, sticky=tk.E)
customer_phone_entry = ttk.Entry(content_frame, width=30)
customer_phone_entry.grid(row=3, column=1, padx=10, pady=5)

customer_score_label = ttk.Label(content_frame, text="Pontuação:")
customer_score_label.grid(row=4, column=0, sticky=tk.E)
customer_score_entry = ttk.Entry(content_frame, width=30)
customer_score_entry.grid(row=4, column=1, padx=10, pady=5)


service_type_combobox.grid(row=0, column=1, padx=10, pady=5)
price_label.grid(row=0, column=2, sticky=tk.E)
price_entry.grid(row=0, column=3, padx=10, pady=5)
customer_name_label.grid(row=1, column=2, sticky=tk.E)
customer_name_entry.grid(row=1, column=3, padx=10, pady=5)
customer_phone_label.grid(row=2, column=2, sticky=tk.E)
customer_phone_entry.grid(row=2, column=3, padx=10, pady=5)
customer_score_label.grid(row=3, column=2, sticky=tk.E)
customer_score_entry.grid(row=3, column=3, padx=10, pady=5)

register_button = ttk.Button(content_frame, text="Registrar", command=register_service)
register_button.grid(row=5, column=1, pady=10)

def update_price(*args):
    service_type = service_type_combobox.get()
    if service_type in service_prices:
        price_value.set(service_prices[service_type])

service_type_combobox.bind("<<ComboboxSelected>>", update_price)

content_frame.columnconfigure(1, weight=1)

root.mainloop()
